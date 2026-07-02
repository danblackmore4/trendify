# Reads the influencer spreadsheet and saves a username → follower count mapping
# to influencers.json. Run this whenever the spreadsheet is updated.

import json
import os
import re
import openpyxl

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Trendify Influencer List.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "influencers.json")

# The spreadsheet has two blank rows at the top before the headers
HEADER_ROW = 3
DATA_START_ROW = 4
USERNAME_COL = "IG Handle"
FOLLOWERS_COL = "Followers"


def parse_followers(value) -> int | None:
    # The spreadsheet stores follower counts as strings like "10.3M" or "249k",
    # so we need to convert them to integers for the engagement score calculation
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    match = re.match(r"^([\d.]+)\s*([MmKk]?)$", text)
    if not match:
        return None
    number, suffix = float(match.group(1)), match.group(2).upper()
    if suffix == "M":
        return int(number * 1_000_000)
    if suffix == "K":
        return int(number * 1_000)
    return int(number)


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Build a dict from column name to 0-based index so we're not hardcoding column letters
    header = {cell.value: cell.column - 1 for cell in ws[HEADER_ROW] if cell.value}
    missing = [c for c in (USERNAME_COL, FOLLOWERS_COL) if c not in header]
    if missing:
        print(f"ERROR: could not find columns {missing} in row {HEADER_ROW}.")
        print(f"Available columns: {list(header.keys())}")
        return

    username_idx = header[USERNAME_COL]
    followers_idx = header[FOLLOWERS_COL]

    influencers = {}
    skipped = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        raw_username = row[username_idx]
        raw_followers = row[followers_idx]

        if not raw_username:
            continue

        # Strip @ prefix if present and normalise to lowercase for consistent lookups
        username = str(raw_username).strip().lstrip("@").lower()
        followers = parse_followers(raw_followers)

        if followers is None:
            print(f"  Warning: could not parse follower count '{raw_followers}' for @{username} — skipping")
            skipped += 1
            continue

        influencers[username] = followers

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(influencers, f, indent=2, ensure_ascii=False)

    print(f"Saved {len(influencers)} influencers to {OUTPUT_FILE}")
    if skipped:
        print(f"Skipped {skipped} rows with unparseable follower counts")


if __name__ == "__main__":
    main()
